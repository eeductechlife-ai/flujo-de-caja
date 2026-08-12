"""
app.py — Servidor Flask para Dashboard Interactivo
Flujo de Caja — Limited Group S.A.
Permite: subir Excel, editar parámetros, generar PDF
"""

import os
import sys
import json
import io
import traceback
from datetime import datetime

try:
    from flask import Flask, render_template, request, jsonify, send_file
    from werkzeug.utils import secure_filename
except ImportError:
    print("❌ Flask no está instalado. Instalando...")
    os.system("pip3 install Flask openpyxl werkzeug --break-system-packages")
    from flask import Flask, render_template, request, jsonify, send_file
    from werkzeug.utils import secure_filename

try:
    import openpyxl
except ImportError:
    os.system("pip3 install openpyxl --break-system-packages")
    import openpyxl

from model import ejecutar_modelo, cargar_config
from pdf_generator import construir_html, exportar_pdf
from excel_loader import extraer_parametros_excel, inyectar_parametros_en_config

# Configuración Flask
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máx
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Estado global (en producción usar DB)
estado_global = {
    'config': None,
    'resultado': None,
    'archivo_cargado': None,
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def cargar_config_default():
    """Carga config.json por defecto."""
    with open('config.json', 'r', encoding='utf-8') as f:
        return json.load(f)


def procesar_excel(archivo_path):
    """Extrae datos de Excel y los mapea a config.json estructura.

    Soporta múltiples formatos de nombres de parámetros del usuario.
    Estructura esperada: Columna A = Parámetro, Columna B = Valor
    """
    wb = openpyxl.load_workbook(archivo_path, data_only=True)

    # Intentar usar hoja "FC Convencional" si existe, sino usar la activa
    if "FC Convencional" in wb.sheetnames:
        ws = wb["FC Convencional"]
    else:
        ws = wb.active

    config = cargar_config_default()

    # Mapeo flexible que reconoce múltiples formas de escribir los parámetros
    mapeos = {
        # OPERACIÓN
        'cantidad anual': ('demanda', 'unidades_base', 'int'),
        'unidades base': ('demanda', 'unidades_base', 'int'),
        'incremento apartir año 4': ('demanda', 'incremento_anno_4', 'decimal'),
        'incremento año 4': ('demanda', 'incremento_anno_4', 'decimal'),
        'precio año 1, 2': ('precios', 'precio_annos_1_2', 'float'),
        'precio años 1-2': ('precios', 'precio_annos_1_2', 'float'),
        'precio apartir del año 3': ('precios', 'precio_anno_3_adelante', 'float'),
        'precio año 3': ('precios', 'precio_anno_3_adelante', 'float'),

        # COSTOS VARIABLES
        'mano de obra': ('costos_variables', 'mano_de_obra', 'float'),
        'materia prima - materiales': ('costos_variables', 'materiales_base', 'float'),
        'materiales año 4 - 6': ('costos_variables', 'materiales_importados', 'float'),
        'materiales importados': ('costos_variables', 'materiales_importados', 'float'),
        'costos indirectos': ('costos_variables', 'costos_indirectos', 'float'),

        # COSTOS FIJOS
        'costo fijos fabricación': ('costos_fijos', 'costo_fijo_fabricacion_base', 'float'),
        'costo fijo fabricación': ('costos_fijos', 'costo_fijo_fabricacion_base', 'float'),
        'costo fijo incremento año 4 - 6': ('costos_fijos', 'incremento_ampliacion', 'float'),
        'gasto admon & venta': ('costos_fijos', 'gastos_admon_ventas_base', 'float'),
        'gasto admon & venta año 4 - 6': ('costos_fijos', 'gastos_admon_ventas_ampliados', 'float'),
        'comisión por venta': ('costos_fijos', 'comision_ventas_pct', 'decimal'),

        # FINANZAS
        'tasa de retorno': ('wacc', None, 'decimal'),
        'wacc': ('wacc', None, 'decimal'),
        'tasa impositiva': ('impuestos', 'tasa_impositiva', 'decimal'),
        'tasa impuesto': ('impuestos', 'tasa_impositiva', 'decimal'),
    }

    procesados = 0
    ignorados = []

    # Iterar filas del Excel
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not row[0] or row[1] is None:
            continue

        clave_raw = str(row[0]).strip()
        valor_raw = row[1]
        clave_norm = clave_raw.lower()

        # Saltar encabezados
        if clave_norm in ['parámetro', 'parameter', 'valor', 'value']:
            continue

        # Buscar coincidencia exacta
        ruta_info = None
        for patron, ruta in mapeos.items():
            if patron == clave_norm:
                ruta_info = ruta
                break

        # Si no hay coincidencia exacta, buscar por substring (más flexible)
        if not ruta_info:
            for patron, ruta in mapeos.items():
                if patron in clave_norm:
                    ruta_info = ruta
                    break

        if not ruta_info:
            ignorados.append(f"'{clave_raw}'")
            continue

        # Extraer ruta y tipo de dato
        ruta, subruta, tipo_dato = ruta_info

        # Convertir valor según tipo
        try:
            if tipo_dato == 'int':
                valor = int(float(valor_raw))
            elif tipo_dato == 'decimal':
                # Si viene como decimal o porcentaje
                valor = float(valor_raw)
                # No convertir si ya es decimal (< 1), solo si es porcentaje
                if valor > 1:
                    valor = valor / 100
            elif tipo_dato == 'float':
                valor = float(valor_raw)
            else:
                valor = float(valor_raw)

            # Guardar en config
            if subruta is None:
                config[ruta] = valor
            else:
                if ruta not in config:
                    config[ruta] = {}
                config[ruta][subruta] = valor

            procesados += 1

        except (ValueError, TypeError) as e:
            ignorados.append(f"'{clave_raw}' (error: {str(e)[:30]})")
            continue

    # Log
    print(f"\n{'='*70}")
    print(f"✓ EXCEL PROCESADO: '{ws.title}'")
    print(f"{'='*70}")
    print(f"✓ Parámetros procesados: {procesados}")
    if ignorados and len(ignorados) <= 5:
        print(f"⚠️  Parámetros ignorados: {', '.join(ignorados)}")
    print(f"{'='*70}\n")

    return config


@app.route('/')
def index():
    """Página principal del dashboard."""
    if estado_global['config'] is None:
        estado_global['config'] = cargar_config_default()
    # ⚠️ NO ejecutar modelo automáticamente - el usuario DEBE cargar un Excel primero
    # El frontend pedirá /api/resultado y obtendrá un error si no hay Excel

    return render_template('index.html')


@app.route('/api/config', methods=['GET'])
def get_config():
    """Retorna la configuración actual en JSON."""
    if estado_global['config'] is None:
        estado_global['config'] = cargar_config_default()
    return jsonify(estado_global['config'])


@app.route('/api/resultado', methods=['GET'])
def get_resultado():
    """Retorna el resultado del modelo (flujo de caja, indicadores, etc)."""
    # 🔧 IMPORTANTE: NO usar fallback a ejecutar_modelo() porque da valores incorrectos en Render
    # El usuario DEBE cargar un Excel explícitamente primero

    if estado_global['resultado'] is None:
        # Intentar cargar desde archivo guardado (para el caso que perista entre requests en la misma sesión)
        if os.path.exists('resultado_excel_cargado.json'):
            try:
                with open('resultado_excel_cargado.json', 'r', encoding='utf-8') as f:
                    estado_global['resultado'] = json.load(f)
                print(f"✅ Resultado cargado desde archivo: resultado_excel_cargado.json")
            except Exception as e:
                print(f"⚠️  Error cargando resultado: {e}")
                # NO ejecutar modelo con valores por defecto; retornar error
                return jsonify({'error': 'Por favor, carga un archivo Excel primero'}), 400
        else:
            # No hay Excel cargado: retornar error en lugar de valores fake
            return jsonify({'error': 'Por favor, carga un archivo Excel primero'}), 400

    resultado = estado_global['resultado']

    # Función auxiliar para convertir valores numéricos
    def convert_to_float(obj):
        if isinstance(obj, dict):
            return {k: convert_to_float(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [convert_to_float(v) if isinstance(v, (int, float, list, dict)) else v for v in obj]
        elif isinstance(obj, (int, float)) and not isinstance(obj, bool):
            return float(obj)
        else:
            return obj

    # Construir respuesta con todos los datos necesarios
    return jsonify({
        'config': resultado['config'],
        'estado_resultados': convert_to_float(resultado['estado_resultados']),
        'flujo_caja': convert_to_float(resultado['flujo_caja']),
        'depreciacion_tabla': convert_to_float(resultado['depreciacion_tabla']),
        'indicadores_sin_vt': convert_to_float(resultado['indicadores_sin_vt']),
        'indicadores_con_vt': convert_to_float(resultado['indicadores_con_vt']),
    })


@app.route('/api/actualizar-parametro', methods=['POST'])
def actualizar_parametro():
    """Actualiza un parámetro y recalcula el modelo."""
    try:
        data = request.json
        ruta = data.get('ruta')  # ej: ['demanda', 'unidades_base']
        valor = data.get('valor')

        if not ruta:
            return jsonify({'error': 'Falta ruta'}), 400

        # Actualizar en config
        config = estado_global['config']
        obj = config
        for clave in ruta[:-1]:
            if clave not in obj:
                obj[clave] = {}
            obj = obj[clave]
        obj[ruta[-1]] = valor

        # Guardar a archivo (backup)
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)

        # Recalcular
        estado_global['resultado'] = ejecutar_modelo('config.json')

        return jsonify({'ok': True, 'mensaje': 'Parámetro actualizado'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/subir-excel', methods=['POST'])
def subir_excel():
    """Recibe archivo Excel y carga VALORES CALCULADOS del flujo de caja."""
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se envió archivo'}), 400

        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'error': 'Archivo vacío'}), 400

        if not allowed_file(archivo.filename):
            return jsonify({'error': 'Solo se permiten .xlsx y .xls'}), 400

        # Guardar temporalmente
        filename = secure_filename(archivo.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        archivo.save(filepath)

        print(f"\n📊 PROCESANDO EXCEL: {filename}")
        print("=" * 70)

        # PASO 1: Extraer parámetros de entrada del Excel
        parametros_excel = extraer_parametros_excel(filepath)
        print(f"✅ Parámetros extraídos: {len(parametros_excel)} grupos")

        # PASO 2: Inyectar parámetros en config.json
        inyectar_parametros_en_config('config.json', parametros_excel)
        estado_global['config'] = cargar_config_default()

        # PASO 3: Extraer el 100% de las tablas del Excel (estado resultados + flujo)
        from excel_loader import extraer_flujo_excel, extraer_estado_flujo_completo
        flujo_excel = extraer_flujo_excel(filepath)
        completo = extraer_estado_flujo_completo(filepath)

        if flujo_excel and flujo_excel.get('flujo_caja'):
            print(f"✅ Flujo de caja extraído del Excel: {len(flujo_excel['flujo_caja'])} períodos")

            # 🔧 100% DESDE EXCEL: aplicar horizonte del Excel al config antes de ejecutar modelo
            horizonte_excel = completo.get('horizonte_evaluacion', 6)
            with open('config.json', 'r', encoding='utf-8') as f:
                _cfg = json.load(f)
            _cfg['horizonte_evaluacion'] = horizonte_excel
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(_cfg, f, indent=2, ensure_ascii=False)
            estado_global['config'] = _cfg
            print(f"✅ Horizonte detectado del Excel: {horizonte_excel} años")

            # 🔧 100% DESDE EXCEL: ejecutar modelo SOLO para la estructura auxiliar
            # (tabla de depreciación por activo, que suma al total del Excel) y config;
            # TODAS las tablas visibles (estado de resultados y flujo de caja) se
            # reemplazan por los valores REALES del Excel, garantizando consistencia.
            config = cargar_config('config.json')
            resultado = ejecutar_modelo('config.json')

            # Estado de Resultados 100% del Excel (reemplazo COMPLETO del dict)
            resultado['estado_resultados'] = dict(completo['estado_resultados'])
            print(f"✅ Estado de Resultados 100% del Excel ({len(completo['estado_resultados'])} filas)")

            # Tabla de Depreciación 100% del Excel (el Excel solo tiene el total)
            resultado['depreciacion_tabla'] = completo['depreciacion_tabla']
            print(f"✅ Tabla de Depreciación 100% del Excel (total real, sin desglose inventado)")

            # Flujo de Caja Libre 100% del Excel (componentes que suman al total)
            resultado['flujo_caja']['utilidad_neta']  = completo['flujo_caja']['utilidad_neta']
            resultado['flujo_caja']['depreciacion']   = completo['flujo_caja']['depreciacion']
            resultado['flujo_caja']['amortizacion']   = completo['flujo_caja']['amortizacion']
            resultado['flujo_caja']['inversiones']    = completo['flujo_caja']['inversiones']
            resultado['flujo_caja']['flujo_kw']       = completo['flujo_caja']['flujo_kw']
            resultado['flujo_caja']['flujo_sin_vt']   = completo['flujo_caja']['flujo_sin_vt']
            resultado['flujo_caja']['flujo_con_vt']   = completo['flujo_caja']['flujo_con_vt']
            resultado['flujo_caja']['valor_terminal'] = completo['flujo_caja']['valor_terminal']
            print(f"✅ Flujo de Caja Libre 100% del Excel (componentes consistentes)")

            # Indicadores (VAN, TIR, WACC, Payback, Índice) 100% del Excel
            if completo.get('indicadores'):
                print(f"✅ Indicadores del Excel:")
                for key, val in completo['indicadores'].items():
                    print(f"   - {key}: {val}")
                    resultado['indicadores_con_vt'][key] = val
                    resultado['indicadores_sin_vt'][key] = val

            # Verificación: imprimir datos actualizados
            print(f"\n✅ VERIFICACIÓN - Datos finales actualizados:")
            if resultado['estado_resultados'].get('ingresos'):
                print(f"   Ingresos (primero): ${resultado['estado_resultados']['ingresos'][0]:,.0f}")
            if resultado['estado_resultados'].get('egresos_op'):
                print(f"   Egresos (primero): ${resultado['estado_resultados']['egresos_op'][0]:,.0f}")
            if resultado['flujo_caja'].get('flujo_con_vt'):
                print(f"   Flujo Caja (primero): ${resultado['flujo_caja']['flujo_con_vt'][0]:,.0f}")
            if resultado['indicadores_con_vt'].get('van'):
                print(f"   VAN: ${resultado['indicadores_con_vt']['van']:,.0f}")
            if resultado['indicadores_con_vt'].get('tir'):
                print(f"   TIR: {resultado['indicadores_con_vt']['tir']*100:.2f}%")

            estado_global['resultado'] = resultado

            # 💾 GUARDAR resultado COMPLETO en archivo para persistencia
            try:
                def _to_serializable(obj):
                    if isinstance(obj, dict):
                        return {k: _to_serializable(v) for k, v in obj.items()}
                    if isinstance(obj, (list, tuple)):
                        return [_to_serializable(v) for v in obj]
                    if isinstance(obj, (int, float, str, bool)) or obj is None:
                        return obj
                    return float(obj) if hasattr(obj, '__float__') else str(obj)

                with open('resultado_excel_cargado.json', 'w', encoding='utf-8') as f:
                    json.dump(_to_serializable(resultado), f, indent=2, ensure_ascii=False)
                print(f"💾 Resultado COMPLETO guardado en: resultado_excel_cargado.json")
            except Exception as e:
                print(f"⚠️  Error guardando resultado: {e}")
        else:
            # Si no hay flujo en Excel, recalcular normalmente
            print(f"⚠️  No se encontró flujo de caja calculado en Excel")
            print(f"   Recalculando con parámetros del Excel...")
            estado_global['resultado'] = ejecutar_modelo('config.json')

        print("=" * 70)
        print(f"✅ Excel procesado exitosamente\n")
        print(f"   Parámetros aplicados: {len(parametros_excel)} grupos")
        if flujo_excel and flujo_excel.get('flujo_caja'):
            print(f"   Valores de flujo cargados del Excel: ✅")
        print("=" * 70 + "\n")

        # Limpiar temp
        os.remove(filepath)

        # Preparar respuesta con datos para debugging
        respuesta = {
            'ok': True,
            'mensaje': f'✅ Excel "{filename}" cargado correctamente',
            'archivo_cargado': filename,
            'parametros_actualizados': len(parametros_excel),
            'flujo_de_excel': bool(flujo_excel and flujo_excel.get('flujo_caja')),
            'debug': {
                'van_en_resultado': estado_global['resultado']['indicadores_con_vt'].get('van'),
                'flujo_primero': estado_global['resultado']['flujo_caja']['flujo_con_vt'][0] if estado_global['resultado']['flujo_caja']['flujo_con_vt'] else None,
                'ingresos_primero': estado_global['resultado']['estado_resultados']['ingresos'][0] if estado_global['resultado']['estado_resultados']['ingresos'] else None
            }
        }

        print(f"\n📤 Retornando a frontend:")
        print(f"   VAN: {respuesta['debug']['van_en_resultado']}")
        print(f"   Flujo: {respuesta['debug']['flujo_primero']}")
        print(f"   Ingresos: {respuesta['debug']['ingresos_primero']}")
        print("=" * 70 + "\n")

        return jsonify(respuesta)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'error': f'Error procesando Excel: {str(e)}'}), 500


@app.route('/api/descargar-pdf', methods=['GET'])
def descargar_pdf():
    """Genera y descarga el PDF del reporte."""
    try:
        if estado_global['resultado'] is None:
            return jsonify({'error': 'Primero ejecuta el modelo'}), 400

        resultado = estado_global['resultado']
        pdf_bytes = exportar_pdf(resultado)

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Flujo_Caja_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/descargar-html', methods=['GET'])
def descargar_html():
    """Genera y descarga el HTML del reporte."""
    try:
        if estado_global['resultado'] is None:
            return jsonify({'error': 'Primero ejecuta el modelo'}), 400

        resultado = estado_global['resultado']
        html = construir_html(resultado)

        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            as_attachment=True,
            download_name=f"Flujo_Caja_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/resetear', methods=['POST'])
def resetear():
    """Resetea a valores por defecto."""
    try:
        estado_global['config'] = cargar_config_default()
        estado_global['resultado'] = ejecutar_modelo('config.json')
        return jsonify({'ok': True, 'mensaje': 'Reseteado a valores por defecto'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    # En la nube, el proveedor asigna el puerto vía la variable de entorno PORT.
    # host=0.0.0.0 permite acceso desde la red local o desde internet (nube).
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    print("\n" + "="*60)
    print("🚀 Flujo de Caja Dashboard")
    print("="*60)
    print(f"📡 Servidor corriendo en http://{host}:{port}")
    print("="*60 + "\n")
    app.run(debug=False, host=host, port=port)
