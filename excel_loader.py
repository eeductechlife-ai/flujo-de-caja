"""
excel_loader.py - Lector robusto de archivos Excel
Extrae TODOS los parámetros y datos del Excel del usuario
"""

import openpyxl
import json
from typing import Dict, Any, Tuple

def extraer_parametros_excel(archivo_path: str) -> Dict[str, Any]:
    """
    Extrae TODOS los parámetros del Excel y mapea a estructura config.json
    """
    wb = openpyxl.load_workbook(archivo_path, data_only=True)

    # Usar la hoja "FC Convencional"
    if "FC Convencional" in wb.sheetnames:
        ws = wb["FC Convencional"]
    else:
        ws = wb.active

    # Mapeo COMPLETO de parámetros del Excel del usuario
    mapeo_excel = {
        # OPERACIÓN
        "cantidad anual (unidad)": ("demanda", "unidades_base", "int"),
        "incremento apartir año 4": ("demanda", "incremento_anno_4", "decimal"),

        # PRECIOS
        "precio año 1, 2": ("precios", "precio_annos_1_2", "float"),
        "precio apartir del año 3": ("precios", "precio_anno_3_adelante", "float"),

        # COSTOS VARIABLES
        "mano de obra": ("costos_variables", "mano_de_obra", "float"),
        "materia prima - materiales": ("costos_variables", "materiales_base", "float"),
        "materiales año 4 - 6": ("costos_variables", "materiales_importados", "float"),
        "costos indirectos": ("costos_variables", "costos_indirectos", "float"),

        # COSTOS FIJOS
        "costo fijos fabricación": ("costos_fijos", "costo_fijo_fabricacion_base", "float"),
        "costo fijo incremento año 4 - 6": ("costos_fijos", "incremento_ampliacion", "float"),
        "gasto admon & venta": ("costos_fijos", "gastos_admon_ventas_base", "float"),
        "gasto admon & venta año 4 - 6": ("costos_fijos", "gastos_admon_ventas_ampliados", "float"),
        "comisión por venta": ("costos_fijos", "comision_ventas_pct", "decimal"),

        # IMPUESTOS Y FINANZAS
        "tasa impositiva": ("impuestos", "tasa_impositiva", "decimal"),
        "tasa de retorno": ("wacc", None, "decimal"),

        # INVERSIONES
        "terreno": ("inversion_inicial", "terreno", "float"),
        "obras fisica inicial": ("inversion_inicial", "obras_fisicas", "float"),
        "maquinaria inicial": ("inversion_inicial", "maquinaria_total", "float"),
        "intangibles": ("inversion_inicial", "intangibles_total", "float"),

        # DEPRECIACIÓN
        "terreno": ("depreciacion", "terreno_annos", "int"),
        "obra física /años": ("depreciacion", "obras_fisicas_annos", "int"),
        "maquinaria/años": ("depreciacion", "maquinaria_annos", "int"),
    }

    parametros = {}

    # Iterar filas del Excel
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row[0] or row[1] is None:
            continue

        clave_raw = str(row[0]).strip().lower()
        valor_raw = row[1]

        # Buscar en mapeo
        if clave_raw not in mapeo_excel:
            continue

        ruta, subruta, tipo = mapeo_excel[clave_raw]

        try:
            # Convertir valor
            if tipo == "int":
                valor = int(float(valor_raw))
            elif tipo == "decimal":
                valor = float(valor_raw)
                if valor > 1:  # Si es porcentaje
                    valor = valor / 100
            elif tipo == "float":
                valor = float(valor_raw)
            else:
                valor = float(valor_raw)

            # Guardar
            if ruta not in parametros:
                parametros[ruta] = {}

            if subruta is None:
                parametros[ruta] = valor
            else:
                parametros[ruta][subruta] = valor

        except (ValueError, TypeError):
            print(f"⚠️  Error procesando: {clave_raw} = {valor_raw}")
            continue

    return parametros


def extraer_flujo_excel(archivo_path: str) -> Dict[str, Any]:
    """
    Extrae el flujo de caja y indicadores del Excel
    """
    wb = openpyxl.load_workbook(archivo_path, data_only=True)

    if "FC Convencional" in wb.sheetnames:
        ws = wb["FC Convencional"]
    else:
        ws = wb.active

    resultados = {
        "años": [],
        "ingresos": [],
        "egresos": [],
        "flujo_caja": [],
        "indicadores": {}
    }

    # Buscar filas de datos
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row[0]:
            continue

        clave = str(row[0]).strip().lower()

        # Extraer años (Fila 25)
        if "periodo/ año" in clave or "año" in clave and row_idx == 25:
            resultados["años"] = [i for i in row[1:] if i is not None and i != "[vacío]"]
            continue

        # Extraer ingresos operativos (Fila 33)
        if "ingreso operativo" in clave:
            resultados["ingresos"] = [float(v) if v else 0 for v in row[1:] if isinstance(v, (int, float))]
            continue

        # Extraer egreso desembolsable (Fila 38)
        if "egreso desembolsable" in clave:
            resultados["egresos"] = [float(v) if v else 0 for v in row[1:] if isinstance(v, (int, float))]
            continue

        # Extraer flujo de caja (Fila 67)
        if "flujo de caja" in clave and row_idx == 67:
            resultados["flujo_caja"] = [float(v) if v else 0 for v in row[1:] if isinstance(v, (int, float))]
            continue

        # Extraer indicadores
        if "vpn" in clave or "van" in clave:
            val = row[1]
            if val and isinstance(val, (int, float)):
                resultados["indicadores"]["van"] = float(val)

        if "tir del proyecto" in clave:
            val = row[1]
            if val and isinstance(val, (int, float)):
                resultados["indicadores"]["tir"] = float(val)

        if "pri (años)" in clave or "periodo de recuperación" in clave:
            val = row[1]
            if val and isinstance(val, (int, float)):
                resultados["indicadores"]["payback"] = float(val)

        if "índice de rentabilidad" in clave:
            val = row[1]
            if val and isinstance(val, (int, float)):
                resultados["indicadores"]["indice_rentabilidad"] = float(val)

        # Extraer WACC / tasa de descuento (celda "WACC (tasa de descuento)" o "Tasa de retorno")
        if ("wacc" in clave or "tasa de descuento" in clave or "tasa de retorno" in clave):
            val = row[1]
            if val and isinstance(val, (int, float)):
                resultados["indicadores"]["wacc"] = float(val)

    return resultados


def extraer_estado_flujo_completo(archivo_path: str) -> Dict[str, Any]:
    """
    Extrae el 100% de las filas de datos del Excel (hoja 'FC Convencional')
    en arrays de 7 elementos alineados a los años 0..6.

    Devuelve las tablas completas (estado de resultados y flujo de caja) tal
    como están en el Excel, para que el dashboard NO recalcule nada.
    """
    wb = openpyxl.load_workbook(archivo_path, data_only=True)
    ws = wb["FC Convencional"] if "FC Convencional" in wb.sheetnames else wb.active

    def fila(n_fila: int, columnas=7) -> list:
        """Lee `columnas` valores desde la columna B (2) de la fila dada."""
        vals = []
        for c in range(2, 2 + columnas):
            v = ws.cell(row=n_fila, column=c).value
            vals.append(float(v) if isinstance(v, (int, float)) else 0.0)
        return vals

    def suma_filas(*filas_n, columnas=7) -> list:
        """Suma varias filas elemento a elemento (para inversiones netas)."""
        acum = [0.0] * columnas
        for nf in filas_n:
            for i, v in enumerate(fila(nf, columnas)):
                acum[i] += v
        return acum

    # ── ESTADO DE RESULTADOS (todas las filas del Excel, hoja FC Convencional) ──
    estado_resultados = {
        "ingresos":       fila(36),   # TOTAL INGRESOS
        "costo_variable": fila(39),   # Costo variable fabricación (negativo)
        "costo_fijo":     fila(40),   # Costo fijos fabricación (negativo)
        "gastos_admon":   fila(41),   # Gastos admon & venta (negativo)
        "comisiones":     fila(42),   # Comisiones por venta (negativo)
        "egresos_op":     fila(38),   # Egreso desembolsable total (negativo)
        "depreciacion":   fila(44),   # Depreciación (negativo)
        "amortizacion":   fila(45),   # Amortización intangible (negativo)
        "valor_libro":    fila(46),   # Valor libro activo (pérdida, negativo)
        "ebit":           fila(49),   # UTILIDAD ANTES DE IMPUESTOS (UAI)
        "impuesto":       fila(50),   # Impuestos (negativo)
        "utilidad_neta":  fila(51),   # UTILIDAD NETA
    }

    # ── FLUJO DE CAJA LIBRE (componentes que suman al flujo, fila 67) ──
    # inversiones netas = inicial(57) + reemplazo(58) + ampliación(59)
    #                     + valor libro(55) + venta activo neto(64)
    flujo_caja = {
        "utilidad_neta": fila(51),                 # UTILIDAD NETA
        "depreciacion":  fila(53),                 # Depreciación (ajuste, positivo)
        "amortizacion":  fila(54),                 # Amortización (ajuste, positivo)
        "inversiones":   suma_filas(57, 58, 59, 55, 64),  # inversiones netas
        "flujo_kw":      fila(60),                 # Inversión/recuperación capital de trabajo
        "flujo_sin_vt":  fila(67),                 # FLUJO DE CAJA (Excel)
        "flujo_con_vt":  fila(67),                 # Igual: el Excel no añade perpetuidad
        "valor_terminal": 0.0,                     # El Excel no incluye valor terminal separado
    }

    # ── TABLA DE DEPRECIACIÓN (el Excel solo tiene el TOTAL, fila 44 y 45) ──
    # Se muestran como positivos. El Excel NO tiene desglose por activo.
    dep_total   = [abs(v) for v in fila(44)]
    amort_total = [abs(v) for v in fila(45)]
    depreciacion_tabla = {
        "Depreciación (del Excel)":  dep_total,
        "Amortización intangibles":  amort_total,
        "Total Depreciación":        [dep_total[i] + amort_total[i] for i in range(len(dep_total))],
    }

    # ── INDICADORES (reutiliza la extracción ya validada) ──
    base = extraer_flujo_excel(archivo_path)

    return {
        "estado_resultados": estado_resultados,
        "flujo_caja": flujo_caja,
        "depreciacion_tabla": depreciacion_tabla,
        "indicadores": base.get("indicadores", {}),
        "flujo_caja_linea": fila(67),  # compatibilidad con extraer_flujo_excel
    }


def inyectar_parametros_en_config(config_path: str, parametros: Dict) -> None:
    """
    Inyecta los parámetros del Excel en el config.json
    """
    import json

    with open(config_path, 'r') as f:
        config = json.load(f)

    # Actualizar config recursivamente
    def actualizar_dict(base, actualizado):
        for clave, valor in actualizado.items():
            if isinstance(valor, dict):
                if clave not in base:
                    base[clave] = {}
                actualizar_dict(base[clave], valor)
            else:
                base[clave] = valor

    actualizar_dict(config, parametros)

    # Guardar
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

    print(f"✅ Config.json actualizado con parámetros del Excel")


if __name__ == "__main__":
    # Test
    archivo = '/Users/home/Desktop/10.FC ACTUALIZADO.xlsx'

    print("Extrayendo parámetros...")
    params = extraer_parametros_excel(archivo)
    print(json.dumps(params, indent=2, default=str))

    print("\n\nExtrayendo flujo...")
    flujo = extraer_flujo_excel(archivo)
    print(json.dumps(flujo, indent=2, default=str))
